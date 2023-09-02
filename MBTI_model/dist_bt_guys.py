# -*- coding: utf-8 -*-
"""
Created on Fri Jul 28 15:05:33 2023

@author: RM275199
"""

import math
import openpyxl as opx

class Profile:
    def __init__(self, name):
        self.E = 0
        self.S = 0
        self.T = 0
        self.J = 0
        self.A = 0
        self.name = name
    
    def import_profile(self, workbook = "stats_personnalities.xlsx"):
        
        file = opx.load_workbook(workbook, data_only=True)
        ws = file['Feuil1']
    
        #ws.cell(20, 14).value = "Test"
        name_line = 4
        while ws.cell(name_line, 2).value.lower() != self.name.lower():
            name_line += 1
        if ws.cell(name_line, 2).value.lower() != self.name.lower():
            print("The name doesn't exist")
        else :
            self.line = name_line
            self.E = ws.cell(self.line, 3).value
            self.S = ws.cell(self.line, 5).value
            self.T = ws.cell(self.line, 7).value
            self.J = ws.cell(self.line, 9).value
            self.A = ws.cell(self.line, 11).value
    
        file.save(workbook)

    
def diff_quadra(a, b):
    return (a-b)**2

def distance_between_2_profiles(p1, p2):
    
    dist  = diff_quadra(p1.E, p2.E)
    dist += diff_quadra(p1.E, p2.E)
    dist += diff_quadra(p1.E, p2.E)
    dist += diff_quadra(p1.E, p2.E)
    dist += diff_quadra(p1.E, p2.E)
    return math.sqrt(dist)


p1 = Profile("Robin")
p1.import_profile()

p2 = Profile("Yigit")
p2.import_profile()

#print(distance_between_2_profiles(p1, p2))


import numpy as np
def test():
    a = np.random.rand(5,1)
    b = np.random.rand(5,1)
    res = np.sqrt(np.sum(np.square(a-b)))

    return res

moy = 0
for _ in range(10000):
    moy += test()

print(moy/10000)
